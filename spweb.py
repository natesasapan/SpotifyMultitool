import os
import streamlit as st
from dotenv import load_dotenv
import spotify_multitool as spm
from openpyxl import Workbook
import io
import json

def main():
    st.title("Spotify Playlist Exporter")

    load_dotenv()

    client_id = os.getenv("CLIENT_ID")
    client_secret = os.getenv("CLIENT_SECRET")
    redirect_uri = os.getenv("URI_REDIR")

    try:
        token = spm.get_token(client_id, client_secret)
        st.success("Connected to Spotify API")
    except Exception as e:
        st.error(f"Error obtaining token: {str(e)}")
        return

    playlist_id = spm.parse_input(st.text_input("Enter Spotify Playlist ID:"))

    if playlist_id and st.button("Export Playlist"):
        try:
            playlist_name = spm.get_playlist_name(token, playlist_id)
            st.write(f"Exporting playlist: {playlist_name}")

            all_tracks = []
            counter = 0
            total_songs = 1

            progress_bar = st.progress(0)

            while counter < total_songs:
                tracks, total_songs = spm.get_songs_from_playlist(token, playlist_id, counter)
                all_tracks.extend(tracks)
                
                counter += 50
                progress = min(counter / total_songs, 1.0)
                progress_bar.progress(progress)

            st.success(f"Retrieved {len(all_tracks)} tracks")

            # Create Excel file
            workbook = Workbook()
            sheet = workbook.active
            for idx, track in enumerate(all_tracks, start=1):
                sheet.cell(row=idx, column=1, value=track['track']['name'])
                sheet.cell(row=idx, column=2, value=track['track']['artists'][0]['name'])
                sheet.cell(row=idx, column=3, value=track['track']['album']['name'])

            excel_file = io.BytesIO()
            workbook.save(excel_file)
            excel_file.seek(0)

            st.download_button(
                label="Download Excel file",
                data=excel_file,
                file_name=f"{playlist_name}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

            # Display first 10 tracks with album covers
            for i in range(min(10, len(all_tracks))):
                track = all_tracks[i]['track']
                image_url = track['album']['images'][0]['url']  # Get the first image URL
                
                col1, col2 = st.columns([1, 3])
                
                with col1:
                    st.image(spm.fetch_album_cover(image_url), width=100)

                with col2:
                    st.write(f"**{track['name']}**")
                    st.write(f"Artist: {track['artists'][0]['name']}")
                    st.write(f"Album: {track['album']['name']}")

                st.write("---")

        except json.JSONDecodeError as e:
            st.error(f"JSON Decode Error: {str(e)}")
            st.error("This error often occurs when the API response is not in the expected format.")
        except Exception as e:
            st.error(f"An error occurred: {str(e)}")
            st.error("Please check your playlist ID and ensure you have the necessary permissions.")

if __name__ == "__main__":
    main()